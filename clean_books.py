"""
clean_books.py
--------------
Reads the parent survey CSV (books_raw or 03_17_2026 format) and produces:

  book_ids_clean.xlsx with two sheets:
    • Book List   — unique books: book_id, book_name (snake_case), raw_variants
    • Child Books — one row per child with book_id_1/2/3 + book_name_1/2/3,
                    ready to merge into child study data by birthday + initial

Deduplication (in order):
  1. Exact normalisation  (lowercase, strip articles/punctuation/whitespace)
  2. Fuzzy clustering     (rapidfuzz token_sort_ratio >= FUZZY_THRESHOLD)
  3. FORCE_SAME overrides — pairs fuzzy misses (e.g. "Good Night Moon" vs "Goodnight Moon")
  4. FORCE_SPLIT overrides — pairs fuzzy wrongly merges (e.g. "Goodnight Moon" vs "Goodnight Mommy")

Usage:
    python clean_books.py <input_csv> [--skip-rows 2]
"""

import re
import argparse
import unicodedata
from datetime import datetime
from collections import defaultdict

import pandas as pd
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

BOOK_COLS = {
    1: ['book_chose_1_1_TEXT', 'book_chose_1_2_TEXT', 'book_chose_1_3_TEXT'],
    2: ['book_chose_1_1_TEXT.1', 'book_chose_1_2_TEXT.1', 'book_chose_1_3_TEXT.1'],
    3: ['book_choice_3_1_TEXT', 'book_choice_3_2_TEXT', 'book_choice_3_3_TEXT'],
}

FUZZY_THRESHOLD = 82

NON_BOOK_RE = re.compile(
    r"can'?t (remember|think|recall)"
    r"|don'?t (remember|recall)"
    r"|we read a variety"
    r"|also can.?t remember"
    r"|^n\/?a$"
    r"|not sure"
    r"|unknown"
    r"|various"
    r"|^(left blank|blank)$",
    re.IGNORECASE,
)

# Pairs to force-merge even if fuzzy score is below threshold.
# Provide normalised keys (lowercase, no articles, no punctuation).
FORCE_SAME: list[tuple[str, str]] = [
    ("good night moon",                        "goodnight moon"),
    ("good night gorilla",                     "goodnight gorilla"),
    ("good night good night construction site", "goodnight goodnight construction site"),
    ("brown bear",                             "brown bear brown bear what do you see"),
    ("brown bear brown bear",                  "brown bear brown bear what do you see"),
    ("knuffle bunny",                          "mr knuffle bunny"),
    ("elephant and piggy",                     "elephant and piggie should i share my ice cream"),
    ("piggy and elephant",                     "elephant and piggie should i share my ice cream"),
    ("elephant piggie series",                 "elephant and piggie should i share my ice cream"),
    ("elephant and piggie all the books",      "elephant and piggie should i share my ice cream"),
    ("elephant piggie goes to a party",        "elephant and piggie should i share my ice cream"),
    ("pet the cat",                            "pete the cat"),
    ("abcs dr seuss",                          "dr seuss abc"),
    ("going to bed book sandra boynton",       "goodnight book by sandra boynton"),
    # "by Eric Carle" suffix shouldn't make a separate book
    ("very hungry caterpillar by eric carle",  "very hungry caterpillar"),
    # Pete the Cat series entries → main Pete the Cat entry
    ("pete the cat series",                    "pete the cat"),
    # "harry potter" generic → book 1
    ("harry potter",                           "harry potter book 1"),
]

# Pairs that should NOT be merged (normalised keys).
FORCE_SPLIT: set[frozenset] = {
    frozenset({"goodnight moon",        "goodnight mommy"}),
    frozenset({"are you my mother",     "you are my heart"}),
    frozenset({"harry potter book 1",   "harry potter book 2"}),
    frozenset({"harry potter book 1",   "harry potter book 3"}),
    frozenset({"harry potter book 2",   "harry potter book 3"}),
    frozenset({"first 100 words",       "first word book"}),
    frozenset({"first 100 words book",  "first word book"}),
    frozenset({"theres a monster in my book", "theres a monster in your book"}),
}

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def parse_bday(s: str):
    s = str(s).strip()
    for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def norm(s: str) -> str:
    """Canonical dedup key: lowercase, no leading article, no punctuation."""
    s = unicodedata.normalize('NFKD', s).lower().strip()
    s = re.sub(r'^(the|a|an)\s+', '', s)
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def to_snake(s: str) -> str:
    """Convert a display string to snake_case."""
    s = unicodedata.normalize('NFKD', s).strip().strip('"\'')
    s = re.sub(r'[^\w]+', '_', s.lower())
    return s.strip('_')


class UF:
    def __init__(self, n):
        self.p = list(range(n))
    def find(self, x):
        while self.p[x] != x:
            self.p[x] = self.p[self.p[x]]
            x = self.p[x]
        return x
    def union(self, x, y):
        self.p[self.find(x)] = self.find(y)
    def same(self, x, y):
        return self.find(x) == self.find(y)


def style_ws(ws, header_color: str):
    hfill = PatternFill('solid', start_color=header_color)
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    alt   = PatternFill('solid', start_color='EEF2F7')
    thin  = Border(bottom=Side(style='thin', color='D0D0D0'))
    for cell in ws[1]:
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20
    for i in range(2, ws.max_row + 1):
        for cell in ws[i]:
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(vertical='center')
            cell.border = thin
            if i % 2 == 0:
                cell.fill = alt


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('input_csv')
    parser.add_argument('--skip-rows', type=int, default=2)
    args = parser.parse_args()

    df = pd.read_csv(args.input_csv, header=0,
                     skiprows=list(range(1, args.skip_rows + 1)))
    print(f"Loaded {len(df)} survey rows")

    # ------------------------------------------------------------------
    # STEP 1 — collect all unique raw book entries
    # ------------------------------------------------------------------
    raw_set: set[str] = set()
    for cols in BOOK_COLS.values():
        for col in cols:
            if col not in df.columns:
                continue
            for val in df[col].dropna():
                val = str(val).strip()
                if val and not NON_BOOK_RE.search(val):
                    raw_set.add(val)

    books = sorted(raw_set)
    norms = [norm(b) for b in books]
    b_idx = {b: i for i, b in enumerate(books)}
    norm_to_idx: dict[str, list[int]] = defaultdict(list)
    for i, n in enumerate(norms):
        norm_to_idx[n].append(i)

    print(f"Unique raw entries: {len(books)}")

    # ------------------------------------------------------------------
    # STEP 2 — cluster
    # ------------------------------------------------------------------
    uf = UF(len(books))

    # 2a. Exact normalisation
    for indices in norm_to_idx.values():
        for j in indices[1:]:
            uf.union(indices[0], j)

    # 2b. Fuzzy (skip force-split pairs)
    for i in range(len(books)):
        for j in range(i + 1, len(books)):
            if uf.same(i, j):
                continue
            ni, nj = norms[i], norms[j]
            if frozenset({ni, nj}) in FORCE_SPLIT:
                continue
            if fuzz.token_sort_ratio(ni, nj) >= FUZZY_THRESHOLD:
                uf.union(i, j)

    # 2c. Force-split: break any bad merges (repeat until stable)
    changed = True
    while changed:
        changed = False
        for fs in FORCE_SPLIT:
            a, b_ = list(fs)
            ai = next((i for i, n in enumerate(norms) if n == a), None)
            bi = next((i for i, n in enumerate(norms) if n == b_), None)
            if ai is None or bi is None:
                continue
            if uf.same(ai, bi):
                uf.p[bi] = bi
                changed = True

    # 2d. Force-same overrides
    for (a_raw, b_raw) in FORCE_SAME:
        na, nb = norm(a_raw), norm(b_raw)
        ai_list = norm_to_idx.get(na, [])
        bi_list = norm_to_idx.get(nb, [])
        if ai_list and bi_list:
            uf.union(ai_list[0], bi_list[0])

    # ------------------------------------------------------------------
    # STEP 3 — build cluster metadata
    # ------------------------------------------------------------------
    raw_counts: dict[str, int] = defaultdict(int)
    for cols in BOOK_COLS.values():
        for col in cols:
            if col not in df.columns:
                continue
            for val in df[col].dropna():
                val = str(val).strip()
                if val and not NON_BOOK_RE.search(val) and val in b_idx:
                    raw_counts[val] += 1

    clusters: dict[int, dict] = {}
    for b in books:
        root = uf.find(b_idx[b])
        if root not in clusters:
            clusters[root] = {'variants': {}}
        clusters[root]['variants'][b] = raw_counts.get(b, 1)

    # Canonical display = most-frequent raw variant
    for root, info in clusters.items():
        info['display'] = max(info['variants'], key=info['variants'].get)

    sorted_roots = sorted(clusters.keys(),
                          key=lambda r: norm(clusters[r]['display']))
    book_id: dict[int, int] = {root: i + 1 for i, root in enumerate(sorted_roots)}

    print(f"Unique books after clustering: {len(clusters)}")

    # ------------------------------------------------------------------
    # STEP 4 — lookup helper: raw string -> (book_id, snake_name)
    # ------------------------------------------------------------------
    def lookup(raw_val: str):
        raw_val = raw_val.strip()
        if raw_val not in b_idx:
            return None, None
        root  = uf.find(b_idx[raw_val])
        bid   = book_id[root]
        snake = to_snake(clusters[root]['display'])
        return bid, snake

    # ------------------------------------------------------------------
    # STEP 5 — build child-level table
    # ------------------------------------------------------------------
    child_rows = []
    for _, row in df.iterrows():
        for child_idx, cols in BOOK_COLS.items():
            name_raw = str(row.get(f'name_{child_idx}', '')).strip()
            bday_raw = str(row.get(f'birthday_{child_idx}', '')).strip()
            if not name_raw or name_raw.lower() == 'nan':
                continue

            name    = name_raw.strip().title()
            bday    = parse_bday(bday_raw)
            initial = name[0].upper() if name else ''

            ids, names = [], []
            for col in cols:
                if col not in df.columns:
                    continue
                val = str(row.get(col, '')).strip()
                if not val or val.lower() == 'nan' or NON_BOOK_RE.search(val):
                    continue
                bid, snake = lookup(val)
                if bid:
                    ids.append(bid)
                    names.append(snake)

            while len(ids) < 3:
                ids.append('')
                names.append('')

            child_rows.append({
                'child_name':    name,
                'first_initial': initial,
                'child_birthday': bday,
                'book_id_1':  ids[0],   'book_name_1': names[0],
                'book_id_2':  ids[1],   'book_name_2': names[1],
                'book_id_3':  ids[2],   'book_name_3': names[2],
            })

    child_df = pd.DataFrame(child_rows)

    # ------------------------------------------------------------------
    # STEP 6 — write Excel
    # ------------------------------------------------------------------
    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Book List'
    ws1.append(['book_id', 'book_name', 'raw_variants'])
    for root in sorted_roots:
        info  = clusters[root]
        bid   = book_id[root]
        snake = to_snake(info['display'])
        variants_str = ' | '.join(sorted(info['variants'].keys(), key=str.lower))
        ws1.append([bid, snake, variants_str])

    style_ws(ws1, '2F5496')
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 48
    ws1.column_dimensions['C'].width = 90

    ws2 = wb.create_sheet('Child Books')
    headers = ['child_name', 'first_initial', 'child_birthday',
               'book_id_1', 'book_name_1',
               'book_id_2', 'book_name_2',
               'book_id_3', 'book_name_3']
    ws2.append(headers)
    for _, r in child_df.iterrows():
        ws2.append([r[h] for h in headers])

    style_ws(ws2, '1A6B3A')
    for col, w in zip('ABCDEFGHI', [20, 10, 14, 10, 38, 10, 38, 10, 38]):
        ws2.column_dimensions[col].width = w

    out = 'book_ids_clean.xlsx'
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"  Book List:   {len(sorted_roots)} unique books")
    print(f"  Child Books: {len(child_df)} child records")


if __name__ == '__main__':
    main()
